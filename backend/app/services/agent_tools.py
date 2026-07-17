from collections.abc import Callable
from dataclasses import dataclass
import base64
import json
import re
from urllib.parse import quote_plus

import httpx

from app.services.curriculum import (
    REFINED_FIELDS,
    create_version_snapshot,
    draft_record,
    load_agent_draft,
    load_document_draft,
    ordered_courses,
    refined_course,
    selected_curriculum_year,
)
from app.services.diffing import diff_course
from app.supabase import first_row, supabase


def _markdown_to_html(md: str) -> str:
    """Convert basic markdown to HTML for PDF generation."""
    html = md
    # Headers
    html = re.sub(r"^### (.+)$", r"<h3>\1</h3>", html, flags=re.MULTILINE)
    html = re.sub(r"^## (.+)$", r"<h2>\1</h2>", html, flags=re.MULTILINE)
    html = re.sub(r"^# (.+)$", r"<h1>\1</h1>", html, flags=re.MULTILINE)
    # Bold
    html = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", html)
    # Italic
    html = re.sub(r"\*(.+?)\*", r"<em>\1</em>", html)
    # Code inline
    html = re.sub(r"`(.+?)`", r"<code>\1</code>", html)
    # Code blocks
    html = re.sub(
        r"```(\w+)?\n(.+?)```",
        r'<pre><code class="language-\1">\2</code></pre>',
        html,
        flags=re.DOTALL,
    )
    # Links
    html = re.sub(r"\[(.+?)\]\((.+?)\)", r'<a href="\2" target="_blank">\1</a>', html)
    # Unordered lists
    html = re.sub(r"^\- (.+)$", r"<li>\1</li>", html, flags=re.MULTILINE)
    html = re.sub(r"(<li>.*?</li>\n)+", r"<ul>\n\g<0></ul>", html)
    # Tables (basic)
    # Split lines
    lines = html.split("\n")
    in_table = False
    result = []
    for line in lines:
        if line.strip().startswith("|") and "|" in line[1:]:
            if not in_table:
                result.append("<table>")
                in_table = True
            cells = [c.strip() for c in line.strip("|").split("|")]
            tag = "th" if not in_table or not result[-1].startswith("<table>") else "td"
            if len(result) == 1 and result[-1] == "<table>":
                tag = "th"
            result.append(
                "<tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>"
            )
        else:
            if in_table:
                result.append("</table>")
                in_table = False
            result.append(line)
    if in_table:
        result.append("</table>")
    html = "\n".join(result)
    # Paragraphs
    lines = html.split("\n")
    result = []
    in_p = False
    for line in lines:
        if line.strip() and not line.startswith("<") and not line.startswith("</"):
            if not in_p:
                result.append("<p>")
                in_p = True
            result.append(line)
        else:
            if in_p:
                result.append("</p>")
                in_p = False
            result.append(line)
    if in_p:
        result.append("</p>")
    return "\n".join(result)


ToolHandler = Callable[[dict], dict]


@dataclass(frozen=True)
class AgentTool:
    name: str
    description: str
    parameters: dict
    handler: ToolHandler

    def schema(self) -> dict:
        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": self.parameters,
            },
        }


def list_tool_schemas() -> list[dict]:
    return [tool.schema() for tool in TOOLS.values()]


def call_tool(name: str, arguments: dict | None = None) -> dict:
    tool = TOOLS.get(name)
    if not tool:
        raise LookupError("Agent tool not found")
    return tool.handler(arguments or {})


def _require_int(arguments: dict, key: str) -> int:
    value = arguments.get(key)
    if value is None:
        raise ValueError(f"{key} is required")
    return int(value)


def _require_dict(arguments: dict, key: str) -> dict:
    value = arguments.get(key)
    if not isinstance(value, dict):
        raise ValueError(f"{key} must be an object")
    return value


def _get_current_course(arguments: dict) -> dict:
    return {"course": refined_course(_require_int(arguments, "refined_id"))}


def _get_course_fields(arguments: dict) -> dict:
    """Get specific field groups from a course. More efficient than full course JSON."""
    refined_id = _require_int(arguments, "refined_id")
    fields = arguments.get("fields")
    if not isinstance(fields, list) or not fields:
        raise ValueError("fields must be a non-empty array of field names")

    course = refined_course(refined_id)
    result = {"refined_id": refined_id}
    for field in fields:
        if field in course:
            result[field] = course[field]
        else:
            result[field] = None
    return result


def _get_course_codes(arguments: dict) -> dict:
    """Get lightweight course identifiers: refined_id, course_code, course_title, semester."""
    refined_id = _require_int(arguments, "refined_id")
    course = refined_course(refined_id)
    return {
        "refined_id": refined_id,
        "course_code": course.get("course_code"),
        "course_title": course.get("course_title"),
        "semester": course.get("semester"),
        "program": course.get("program"),
    }


def _get_course_syllabus(arguments: dict) -> dict:
    """Get syllabus content: units, objectives, course_outcomes."""
    refined_id = _require_int(arguments, "refined_id")
    course = refined_course(refined_id)
    return {
        "refined_id": refined_id,
        "units": course.get("units"),
        "objectives": course.get("objectives"),
        "course_outcomes": course.get("course_outcomes"),
    }


def _get_course_textbooks(arguments: dict) -> dict:
    """Get textbook fields: text_books, reference_books."""
    refined_id = _require_int(arguments, "refined_id")
    course = refined_course(refined_id)
    return {
        "refined_id": refined_id,
        "text_books": course.get("text_books"),
        "reference_books": course.get("reference_books"),
    }


def _get_course_deterministic(arguments: dict) -> dict:
    """Get deterministic fields (program, hours, credits, course_type) — these are agent-protected."""
    refined_id = _require_int(arguments, "refined_id")
    course = refined_course(refined_id)
    return {
        "refined_id": refined_id,
        "program": course.get("program"),
        "lecture_hours": course.get("lecture_hours"),
        "tutorial_hours": course.get("tutorial_hours"),
        "practical_hours": course.get("practical_hours"),
        "self_study": course.get("self_study"),
        "credits": course.get("credits"),
        "course_type": course.get("course_type"),
    }


def _get_course_lab(arguments: dict) -> dict:
    """Get lab experiments and tools/languages."""
    refined_id = _require_int(arguments, "refined_id")
    course = refined_course(refined_id)
    return {
        "refined_id": refined_id,
        "lab_experiments": course.get("lab_experiments"),
        "tools_languages": course.get("tools_languages"),
    }


def _diff_course_json(arguments: dict) -> dict:
    return diff_course(
        _require_dict(arguments, "current"), _require_dict(arguments, "proposed")
    )


def _create_course_draft(arguments: dict) -> dict:
    refined_id = _require_int(arguments, "refined_id")
    if refined_id <= 0:
        raise ValueError(
            "refined_id must be a valid existing course ID. To create a brand-new course, use create_refined_course instead."
        )
    fields = arguments.get("fields")
    if not isinstance(fields, dict) or not fields:
        raise ValueError(
            'fields must be a non-empty object containing only the fields to change, e.g. {"text_books": "new value"}. Do not pass all course data; only pass what should change.'
        )
    record = draft_record(refined_id, fields, str(arguments.get("reason") or ""))
    draft = supabase.table("agent_drafts").insert(record).execute().data[0]
    return {"draft": draft}


_ARRAY_FIELDS = {
    "course_outcomes",
    "lab_experiments",
    "objectives",
    "text_books",
    "reference_books",
    "units",
}


def _coerce_array(value):
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("["):
            try:
                parsed = json.loads(stripped)
                if isinstance(parsed, list):
                    return parsed
            except json.JSONDecodeError:
                pass
        return [stripped] if stripped else []
    return [value] if value is not None else []


def _create_refined_course(arguments: dict) -> dict:
    from app.services.deterministic import (
        compute_course_type,
        compute_hours,
        compute_program,
    )

    credit_category = str(arguments.get("credit_category") or "4")
    target_dept = str(arguments.get("target_department") or "CSE")
    det = compute_hours(credit_category)

    computed = {
        "program": compute_program(target_dept),
        "course_type": compute_course_type(credit_category),
        "status": "draft",
    }
    for key in (
        "lecture_hours",
        "tutorial_hours",
        "practical_hours",
        "self_study",
        "credits",
    ):
        if key in arguments:
            computed[key] = int(arguments[key] or 0)
        else:
            computed[key] = det[key]
    if "semester" in arguments:
        computed["semester"] = int(arguments["semester"] or 0)

    fields = {
        k: v for k, v in arguments.items() if v is not None and k in REFINED_FIELDS
    }
    for key in _ARRAY_FIELDS:
        if key in fields:
            fields[key] = _coerce_array(fields[key])
    fields.update(computed)

    refined_id = arguments.get("refined_id")
    if refined_id:
        result = (
            supabase.table("refined_submissions")
            .update(fields)
            .eq("id", int(refined_id))
            .execute()
        )
        row = result.data[0] if result.data else None
        return {"refined_id": int(refined_id), "updated": True, "course": row}

    if "submission_id" not in fields or not fields.get("submission_id"):
        placeholder = (
            supabase.table("submissions")
            .insert(
                {
                    "faculty_email": "ai-generated@pes.edu",
                    "course_title": fields.get("course_title") or "",
                    "offering_department": "CS",
                    "target_department": arguments.get("target_department") or "CSE",
                    "semester": int(arguments.get("semester") or 1),
                    "credit_category": arguments.get("credit_category") or "4",
                    "raw_course_content": fields.get("prelude") or "AI-created course",
                    "text_books": "",
                    "reference_books": "",
                    "preferred_tools": "",
                    "status": "refined",
                }
            )
            .execute()
            .data[0]
        )
        fields["submission_id"] = placeholder["id"]

    result = supabase.table("refined_submissions").insert(fields).execute()
    row = result.data[0]
    return {"refined_id": row["id"], "updated": False, "course": row}


def _get_curriculum_json(arguments: dict) -> dict:
    query = supabase.table("refined_submissions").select("*").in_("status", ["refined"])
    if arguments.get("semester") is not None:
        query = query.eq("semester", int(arguments["semester"]))
    return {"courses": ordered_courses(query.execute().data)}


def _create_document_draft(arguments: dict) -> dict:
    courses = arguments.get("courses")
    if not isinstance(courses, list) or not courses:
        raise ValueError("courses must be a non-empty array")

    records = []
    for course in courses:
        if not isinstance(course, dict):
            raise ValueError("each course must be an object")
        records.append(
            draft_record(
                int(course.get("refined_id")),
                _require_dict(course, "fields"),
                str(arguments.get("reason") or ""),
            )
        )

    summaries = [record["diff_summary"] for record in records]
    document_summary = {
        "courses_changed": len(records),
        "courses_with_removed_topics": sum(
            1 for summary in summaries if summary.get("topics_removed")
        ),
        "courses_with_protected_changes": sum(
            1 for summary in summaries if summary.get("protected_changes")
        ),
        "max_syllabus_change_percent": max(
            (summary.get("syllabus_change_percent") or 0 for summary in summaries),
            default=0,
        ),
    }
    document = (
        supabase.table("agent_document_drafts")
        .insert(
            {
                "curriculum_version_id": arguments.get("curriculum_version_id"),
                "uploaded_document_id": str(
                    arguments.get("uploaded_document_id") or ""
                ).strip(),
                "diff_summary": document_summary,
                "change_reason": str(arguments.get("reason") or "").strip(),
                "status": "blocked"
                if document_summary["courses_with_protected_changes"]
                else "proposed",
            }
        )
        .execute()
        .data[0]
    )
    for record in records:
        record["document_draft_id"] = document["id"]
    drafts = supabase.table("agent_drafts").insert(records).execute().data
    return {"document_draft": document, "drafts": drafts}


def _get_course_draft(arguments: dict) -> dict:
    return {"draft": load_agent_draft(_require_int(arguments, "draft_id"))}


def _get_document_draft(arguments: dict) -> dict:
    return load_document_draft(_require_int(arguments, "document_draft_id"))


def _get_preview_url(arguments: dict) -> dict:
    kind = str(arguments.get("kind") or "")
    item_id = _require_int(arguments, "id")
    paths = {
        "course": f"/api/preview/course/{item_id}",
        "draft": f"/api/agent/drafts/{item_id}/preview",
        "document_draft": f"/api/agent/document-drafts/{item_id}/preview",
    }
    if kind not in paths:
        raise ValueError("kind must be course, draft, or document_draft")
    return {"url": paths[kind]}


def _list_courses(arguments: dict) -> dict:
    query = (
        supabase.table("refined_submissions")
        .select("id,semester,course_code,course_title")
        .neq("status", "archived")
    )
    if arguments.get("semester") is not None:
        query = query.eq("semester", int(arguments["semester"]))
    rows = query.execute().data
    rows.sort(
        key=lambda row: (
            int(row.get("semester") or 0),
            str(row.get("course_code") or ""),
            int(row.get("id") or 0),
        )
    )
    return {"courses": rows}


def _fetch_url(arguments: dict) -> dict:
    url = str(arguments.get("url") or "").strip()
    if not url:
        raise ValueError("url is required")
    resp = httpx.get(url, timeout=30, follow_redirects=True)
    resp.raise_for_status()
    text = resp.text[:15000]
    return {"url": url, "text": text, "chars": len(text)}


def _web_search(arguments: dict) -> dict:
    """Search the web using DuckDuckGo's HTML endpoint and return top results."""
    query = str(arguments.get("query") or "").strip()
    if not query:
        raise ValueError("query is required")
    num_results = int(arguments.get("num_results") or 5)
    num_results = min(max(num_results, 1), 10)

    url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}"
    resp = httpx.get(
        url, timeout=15, follow_redirects=True, headers={"User-Agent": "Mozilla/5.0"}
    )
    resp.raise_for_status()

    # Parse results from DuckDuckGo HTML
    html = resp.text
    results = []
    # Pattern for result snippets
    result_pattern = re.compile(r'class="result__snippet">(.*?)</a>', re.DOTALL)
    title_pattern = re.compile(r'class="result__title">.*?>(.*?)</a>', re.DOTALL)
    url_pattern = re.compile(r'class="result__url">.*?>(.*?)</a>', re.DOTALL)

    snippets = result_pattern.findall(html)
    titles = title_pattern.findall(html)
    urls = url_pattern.findall(html)

    for i in range(min(num_results, len(snippets))):
        title = re.sub(r"<[^>]+>", "", titles[i] if i < len(titles) else "").strip()
        snippet = re.sub(r"<[^>]+>", "", snippets[i]).strip()
        link = re.sub(r"<[^>]+>", "", urls[i] if i < len(urls) else "").strip()
        if title or snippet:
            results.append({"title": title, "snippet": snippet[:300], "url": link})

    return {"query": query, "results": results}


def _create_report(arguments: dict) -> dict:
    session_id = _require_int(arguments, "session_id")
    content = str(arguments.get("content") or "").strip()
    if not content:
        raise ValueError("content is required")
    filename = str(arguments.get("filename") or "report.md").strip()
    fmt = str(arguments.get("format") or "markdown").strip().lower()
    if fmt not in ("markdown", "pdf"):
        raise ValueError("format must be 'markdown' or 'pdf'")

    if fmt == "pdf":
        # Convert markdown to HTML then to PDF
        # Simple markdown to HTML conversion
        html_content = _markdown_to_html(content)
        html_full = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; max-width: 800px; margin: 0 auto; padding: 20px; }}
        h1, h2, h3 {{ color: #00377b; }}
        table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background: #f0f2f5; }}
        code {{ background: #f3f4f6; padding: 2px 4px; border-radius: 3px; }}
        pre {{ background: #f3f4f6; padding: 10px; overflow-x: auto; }}
    </style>
</head>
<body>
{html_content}
</body>
</html>"""

        # Use weasyprint to generate PDF
        from weasyprint import HTML

        pdf_bytes = HTML(string=html_full, base_url=".").write_pdf()

        if filename.endswith(".md"):
            filename = filename[:-3] + ".pdf"
        elif not filename.endswith(".pdf"):
            filename = filename + ".pdf"

        content_type = "application/pdf"
        size_bytes = len(pdf_bytes)
        extracted_text = f"[PDF file - {size_bytes} bytes]"
        content_base64 = base64.b64encode(pdf_bytes).decode()
    else:
        pdf_bytes = None
        content_type = "text/markdown"
        size_bytes = len(content.encode())
        extracted_text = content
        content_base64 = ""

    row = (
        supabase.table("chat_attachments")
        .insert(
            {
                "session_id": session_id,
                "filename": filename,
                "content_type": content_type,
                "size_bytes": size_bytes,
                "extracted_text": extracted_text,
                "content_base64": content_base64,
                "status": "ready",
            }
        )
        .execute()
        .data[0]
    )

    return {
        "attachment": {
            "id": row["id"],
            "filename": row["filename"],
            "chars": size_bytes,
            "format": fmt,
        }
    }


def _attachment_text(arguments: dict) -> dict:
    session_id = _require_int(arguments, "session_id")
    ids = [int(value) for value in arguments.get("attachment_ids") or []]
    if not ids:
        raise ValueError("attachment_ids is required")
    rows = (
        supabase.table("chat_attachments")
        .select("id,filename,status,error,extracted_text")
        .eq("session_id", session_id)
        .in_("id", ids)
        .execute()
        .data
    )
    return {"attachments": rows}


def _create_curriculum_version(arguments: dict) -> dict:
    name = str(arguments.get("name") or "").strip()
    if not name:
        raise ValueError("name is required")
    version = create_version_snapshot(name)
    return {"version": version}


def _define_specialization(arguments: dict) -> dict:
    semester = _require_int(arguments, "semester")
    name = str(arguments.get("name") or "").strip()
    if not name:
        raise ValueError("name is required")
    letter = str(arguments.get("letter") or "").strip().upper()
    if not letter:
        existing = (
            supabase.table("specialization_definitions")
            .select("letter")
            .eq("semester", semester)
            .execute()
            .data
        )
        used = {row["letter"] for row in existing}
        for candidate in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            if candidate not in used:
                letter = candidate
                break
        if not letter:
            raise ValueError(
                "No free specialization letter available for this semester"
            )
    academic_year = str(arguments.get("academic_year") or "").strip()
    key = str(arguments.get("key") or "").strip().upper()
    if not key:
        key = name.split("(")[-1].rstrip(")") if "(" in name else name[:3].upper()
    row = (
        supabase.table("specialization_definitions")
        .insert(
            {
                "semester": semester,
                "letter": letter,
                "name": name,
                "key": key,
                "academic_year": academic_year,
            }
        )
        .execute()
        .data[0]
    )
    return {"specialization": row}


def _list_specializations(arguments: dict) -> dict:
    query = supabase.table("specialization_definitions").select("*")
    if arguments.get("semester") is not None:
        query = query.eq("semester", int(arguments["semester"]))
    return {"specializations": query.order("semester").order("letter").execute().data}


def _assign_elective_to_tracks(arguments: dict) -> dict:
    refined_id = _require_int(arguments, "refined_id")
    spec_ids = arguments.get("specialization_ids")
    if not isinstance(spec_ids, list) or not spec_ids:
        raise ValueError("specialization_ids must be a non-empty array")
    created = 0
    for spec_id in spec_ids:
        existing = (
            supabase.table("course_specialization_assignments")
            .select("id")
            .eq("refined_id", refined_id)
            .eq("specialization_id", int(spec_id))
            .execute()
            .data
        )
        if not existing:
            supabase.table("course_specialization_assignments").insert(
                {"refined_id": refined_id, "specialization_id": int(spec_id)}
            ).execute()
            created += 1
    supabase.table("refined_submissions").update({"is_elective": True}).eq(
        "id", refined_id
    ).execute()
    return {"assignments_created": created}


def _remove_elective_from_tracks(arguments: dict) -> dict:
    refined_id = _require_int(arguments, "refined_id")
    spec_ids = arguments.get("specialization_ids")
    if not isinstance(spec_ids, list) or not spec_ids:
        raise ValueError("specialization_ids must be a non-empty array")
    removed = 0
    for spec_id in spec_ids:
        result = (
            supabase.table("course_specialization_assignments")
            .delete()
            .eq("refined_id", refined_id)
            .eq("specialization_id", int(spec_id))
            .execute()
        )
        removed += len(result.data or [])
    return {"assignments_removed": removed}


def _get_course_assignments(arguments: dict) -> dict:
    refined_id = _require_int(arguments, "refined_id")
    assignments = (
        supabase.table("course_specialization_assignments")
        .select("*, specialization_definitions(*)")
        .eq("refined_id", refined_id)
        .execute()
        .data
    )
    return {"refined_id": refined_id, "assignments": assignments}


def _update_deterministic_fields(arguments: dict) -> dict:
    from app.services.diffing import PROTECTED_FIELDS

    refined_id = _require_int(arguments, "refined_id")
    fields = _require_dict(arguments, "fields")
    protected = {key: value for key, value in fields.items() if key in PROTECTED_FIELDS}
    if not protected:
        raise ValueError(
            "No deterministic fields provided. Use create_course_draft for other fields."
        )
    record = draft_record(
        refined_id,
        protected,
        str(
            arguments.get("reason")
            or "Explicit user request to change deterministic fields"
        ),
    )
    record["status"] = "blocked"
    draft = supabase.table("agent_drafts").insert(record).execute().data[0]
    return {
        "draft": draft,
        "warning": "This draft modifies deterministic fields (program, hours, credits, course_type). The user must explicitly approve it in the Review panel before it is applied.",
    }


def _signal_done(arguments: dict) -> dict:
    summary = str(arguments.get("summary") or "").strip()
    if not summary:
        raise ValueError("summary is required")
    return {"done": True, "summary": summary}


def _create_spreadsheet(arguments: dict) -> dict:
    """Generate CSV or Excel file from rows data and save as chat attachment."""
    session_id = _require_int(arguments, "session_id")
    rows = arguments.get("rows")
    if not isinstance(rows, list) or not rows:
        raise ValueError("rows must be a non-empty array of objects")
    columns = arguments.get("columns")
    if not isinstance(columns, list) or not columns:
        raise ValueError("columns must be a non-empty array of column names")
    filename = str(arguments.get("filename") or "spreadsheet.csv").strip()
    fmt = str(arguments.get("format") or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        raise ValueError("format must be 'csv' or 'xlsx'")

    if fmt == "xlsx":
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="00377B", end_color="00377B", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row_idx in range(2, len(rows) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, min(len(val), 60))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
                max_len + 2
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        file_bytes = buf.read()

        if not filename.endswith(".xlsx"):
            filename = filename.rsplit(".", 1)[0] + ".xlsx"
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        extracted_text = f"[Excel file - {len(file_bytes)} bytes, {len(rows)} rows]"
        content_base64 = base64.b64encode(file_bytes).decode()
        size_bytes = len(file_bytes)
    else:
        import io
        import csv

        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row_data in rows:
            writer.writerow({k: v for k, v in row_data.items() if k in columns})
        csv_text = buf.getvalue()

        if not filename.endswith(".csv"):
            filename = filename.rsplit(".", 1)[0] + ".csv"
        content_type = "text/csv"
        extracted_text = csv_text
        content_base64 = ""
        size_bytes = len(csv_text.encode())

    row = (
        supabase.table("chat_attachments")
        .insert(
            {
                "session_id": session_id,
                "filename": filename,
                "content_type": content_type,
                "size_bytes": size_bytes,
                "extracted_text": extracted_text,
                "content_base64": content_base64,
                "status": "ready",
            }
        )
        .execute()
        .data[0]
    )
    return {
        "attachment": {
            "id": row["id"],
            "filename": row["filename"],
            "rows": len(rows),
            "columns": columns,
            "format": fmt,
        }
    }


def _get_version(arguments: dict) -> dict:
    """Load a curriculum version snapshot with its course list."""
    version_id = _require_int(arguments, "version_id")
    version_row = first_row(
        supabase.table("curriculum_versions").select("*").eq("id", version_id)
    )
    if not version_row:
        raise ValueError(f"Version {version_id} not found")
    snapshot_rows = (
        supabase.table("finalized_submissions")
        .select("refined_id,course_json")
        .eq("curriculum_version_id", version_id)
        .order("refined_id")
        .execute()
        .data
    )
    courses = []
    for snap in snapshot_rows:
        cj = snap.get("course_json") or {}
        courses.append(
            {
                "refined_id": snap.get("refined_id"),
                "course_code": cj.get("course_code", ""),
                "course_title": cj.get("course_title", ""),
                "semester": cj.get("semester", ""),
                "credits": cj.get("credits", ""),
            }
        )
    return {"version": version_row, "courses": courses, "course_count": len(courses)}


def _diff_versions(arguments: dict) -> dict:
    """Compare two curriculum version snapshots. Shows added, removed, and changed courses."""
    version_a_id = _require_int(arguments, "version_a")
    version_b_id = _require_int(arguments, "version_b")

    def _load_snapshot(vid: int) -> dict:
        rows = (
            supabase.table("finalized_submissions")
            .select("refined_id,course_json")
            .eq("curriculum_version_id", vid)
            .execute()
            .data
        )
        return {row["refined_id"]: row.get("course_json") or {} for row in rows}

    snap_a = _load_snapshot(version_a_id)
    snap_b = _load_snapshot(version_b_id)

    all_ids = set(snap_a.keys()) | set(snap_b.keys())
    added = []
    removed = []
    changed = []
    unchanged = []

    for rid in sorted(all_ids):
        a = snap_a.get(rid)
        b = snap_b.get(rid)
        if a and not b:
            removed.append(
                {
                    "refined_id": rid,
                    "course_code": a.get("course_code", ""),
                    "course_title": a.get("course_title", ""),
                }
            )
        elif b and not a:
            added.append(
                {
                    "refined_id": rid,
                    "course_code": b.get("course_code", ""),
                    "course_title": b.get("course_title", ""),
                }
            )
        elif a != b:
            d = diff_course(a, b)
            changed.append(
                {
                    "refined_id": rid,
                    "course_code": b.get("course_code", ""),
                    "course_title": b.get("course_title", ""),
                    "change_percent": d.get("change_percent", 0),
                    "syllabus_change_percent": d.get("syllabus_change_percent", 0),
                    "topics_added": d.get("topics_added", []),
                    "topics_removed": d.get("topics_removed", []),
                }
            )
        else:
            unchanged.append(rid)

    return {
        "version_a": version_a_id,
        "version_b": version_b_id,
        "added": added,
        "removed": removed,
        "changed": changed,
        "unchanged_count": len(unchanged),
        "summary": {
            "added": len(added),
            "removed": len(removed),
            "changed": len(changed),
            "unchanged": len(unchanged),
        },
    }


def _get_curriculum_stats(arguments: dict) -> dict:
    """Compute aggregate statistics for the curriculum or a specific semester."""
    query = (
        supabase.table("refined_submissions")
        .select(
            "semester,credits,course_type,credit_category,lecture_hours,practical_hours,visible,is_elective"
        )
        .in_("status", ["refined"])
    )
    if arguments.get("semester") is not None:
        query = query.eq("semester", int(arguments["semester"]))
    rows = query.execute().data

    if not rows:
        return {"stats": {}, "message": "No courses found"}

    by_semester: dict[int, dict] = {}
    for row in rows:
        sem = int(row.get("semester") or 0)
        if sem not in by_semester:
            by_semester[sem] = {
                "total": 0,
                "visible": 0,
                "total_credits": 0,
                "electives": 0,
                "course_types": {},
                "credit_categories": {},
            }
        s = by_semester[sem]
        s["total"] += 1
        if row.get("visible", True):
            s["visible"] += 1
        s["total_credits"] += int(row.get("credits") or 0)
        if row.get("is_elective"):
            s["electives"] += 1
        ct = str(row.get("course_type") or "Unknown")
        s["course_types"][ct] = s["course_types"].get(ct, 0) + 1
        cc = str(row.get("credit_category") or "?")
        s["credit_categories"][cc] = s["credit_categories"].get(cc, 0) + 1

    total_credits = sum(s["total_credits"] for s in by_semester.values())
    total_courses = sum(s["total"] for s in by_semester.values())

    return {
        "total_courses": total_courses,
        "total_credits": total_credits,
        "by_semester": {str(k): v for k, v in sorted(by_semester.items())},
    }


def _batch_read_courses(arguments: dict) -> dict:
    """Read specific fields from multiple courses in one call. More efficient than multiple get_course_fields calls."""
    refined_ids = arguments.get("refined_ids")
    if not isinstance(refined_ids, list) or not refined_ids:
        raise ValueError("refined_ids must be a non-empty array of integers")
    fields = arguments.get("fields")
    if not isinstance(fields, list) or not fields:
        raise ValueError("fields must be a non-empty array of field names")

    results = []
    for rid in refined_ids:
        rid = int(rid)
        course = refined_course(rid)
        if course is None:
            results.append({"refined_id": rid, "error": "Course not found"})
            continue
        entry = {"refined_id": rid}
        for field in fields:
            entry[field] = course.get(field)
        results.append(entry)

    return {"courses": results, "count": len(results)}


OBJECT = {"type": "object", "additionalProperties": False}

TOOLS: dict[str, AgentTool] = {
    "get_current_course_json": AgentTool(
        "get_current_course_json",
        "Read the current template-ready JSON for one refined course.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_current_course,
    ),
    "get_course_codes": AgentTool(
        "get_course_codes",
        "Read lightweight course identifiers (refined_id, course_code, course_title, semester, program). Use for listing or quick lookups.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_codes,
    ),
    "get_course_syllabus": AgentTool(
        "get_course_syllabus",
        "Read syllabus content: units, objectives, course_outcomes.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_syllabus,
    ),
    "get_course_textbooks": AgentTool(
        "get_course_textbooks",
        "Read textbook fields: text_books, reference_books.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_textbooks,
    ),
    "get_course_deterministic": AgentTool(
        "get_course_deterministic",
        "Read deterministic/protected fields: program, lecture_hours, tutorial_hours, practical_hours, self_study, credits, course_type. These cannot be changed by the agent.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_deterministic,
    ),
    "get_course_lab": AgentTool(
        "get_course_lab",
        "Read lab experiments and tools/languages.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_lab,
    ),
    "get_course_fields": AgentTool(
        "get_course_fields",
        "Read arbitrary specific fields from a course. Provide a list of field names. More efficient than fetching full JSON when you only need a subset.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {"type": "integer"},
                "fields": {"type": "array", "items": {"type": "string"}, "minItems": 1},
            },
            "required": ["refined_id", "fields"],
        },
        _get_course_fields,
    ),
    "diff_course_json": AgentTool(
        "diff_course_json",
        "Compare two course JSON objects and return patch operations, changed percent, and syllabus topic changes.",
        {
            **OBJECT,
            "properties": {
                "current": {"type": "object"},
                "proposed": {"type": "object"},
            },
            "required": ["current", "proposed"],
        },
        _diff_course_json,
    ),
    "create_course_draft": AgentTool(
        "create_course_draft",
        "Create a human-reviewable draft for one course. This never applies changes to refined_submissions.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {"type": "integer"},
                "fields": {"type": "object"},
                "reason": {"type": "string"},
            },
            "required": ["refined_id", "fields"],
        },
        _create_course_draft,
    ),
    "create_refined_course": AgentTool(
        "create_refined_course",
        "Create a new course directly in refined_submissions, or update an existing one by refined_id. Use this ONLY for brand-new courses that do not exist in the curriculum yet. For modifications to existing courses, use create_course_draft instead. Call signal_done immediately after this tool. Do not retry if this tool returns an error; report the error to the user instead.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {
                    "type": "integer",
                    "description": "If provided, update this existing course. If omitted, create a new one.",
                },
                "course_code": {"type": "string", "description": "e.g. UE25CS353A"},
                "course_title": {"type": "string"},
                "semester": {"type": "integer", "minimum": 1, "maximum": 8},
                "target_department": {
                    "type": "string",
                    "description": "CSE, ECE, ME, BT, EEE, AIML",
                },
                "credit_category": {"type": "string", "description": "0, 2, 4, or 5"},
                "units": {
                    "type": "array",
                    "description": "Course units. Each unit must have unit_number (int), title (str), content (str, the syllabus text), and hours (int).",
                    "items": {
                        "type": "object",
                        "properties": {
                            "unit_number": {"type": "integer"},
                            "title": {"type": "string"},
                            "content": {"type": "string"},
                            "hours": {"type": "integer"},
                        },
                        "required": ["unit_number", "title", "content", "hours"],
                    },
                },
                "objectives": {"type": "string"},
                "course_outcomes": {"type": "string"},
                "text_books": {"type": "string"},
                "reference_books": {"type": "string"},
                "lab_experiments": {"type": "string"},
                "tools_languages": {"type": "string"},
                "prelude": {"type": "string"},
                "desirable_knowledge": {"type": "string"},
            },
            "required": [
                "course_code",
                "course_title",
                "semester",
                "target_department",
                "credit_category",
            ],
        },
        _create_refined_course,
    ),
    "get_curriculum_json": AgentTool(
        "get_curriculum_json",
        "Read template-ready JSON for the full curriculum, optionally filtered by semester.",
        {
            **OBJECT,
            "properties": {"semester": {"type": "integer", "minimum": 1, "maximum": 8}},
        },
        _get_curriculum_json,
    ),
    "create_document_draft": AgentTool(
        "create_document_draft",
        "Create one human-reviewable document draft containing proposed changes for multiple courses. This never applies changes.",
        {
            **OBJECT,
            "properties": {
                "courses": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "refined_id": {"type": "integer"},
                            "fields": {"type": "object"},
                        },
                        "required": ["refined_id", "fields"],
                        "additionalProperties": False,
                    },
                },
                "reason": {"type": "string"},
                "uploaded_document_id": {"type": "string"},
                "curriculum_version_id": {"type": "integer"},
            },
            "required": ["courses"],
        },
        _create_document_draft,
    ),
    "get_course_draft": AgentTool(
        "get_course_draft",
        "Read one staged course draft and its diff summary.",
        {
            **OBJECT,
            "properties": {"draft_id": {"type": "integer"}},
            "required": ["draft_id"],
        },
        _get_course_draft,
    ),
    "get_document_draft": AgentTool(
        "get_document_draft",
        "Read one staged document draft and all linked course drafts.",
        {
            **OBJECT,
            "properties": {"document_draft_id": {"type": "integer"}},
            "required": ["document_draft_id"],
        },
        _get_document_draft,
    ),
    "get_preview_url": AgentTool(
        "get_preview_url",
        "Return the preview URL for a course, course draft, or document draft.",
        {
            **OBJECT,
            "properties": {
                "kind": {
                    "type": "string",
                    "enum": ["course", "draft", "document_draft"],
                },
                "id": {"type": "integer"},
            },
            "required": ["kind", "id"],
        },
        _get_preview_url,
    ),
    "list_courses": AgentTool(
        "list_courses",
        "List refined course IDs and titles, optionally filtered by semester.",
        {
            **OBJECT,
            "properties": {"semester": {"type": "integer", "minimum": 1, "maximum": 8}},
        },
        _list_courses,
    ),
    "get_attachment_text": AgentTool(
        "get_attachment_text",
        "Read extracted text for uploaded chat attachments within a chat session.",
        {
            **OBJECT,
            "properties": {
                "session_id": {"type": "integer"},
                "attachment_ids": {"type": "array", "items": {"type": "integer"}},
            },
            "required": ["session_id", "attachment_ids"],
        },
        _attachment_text,
    ),
    "fetch_url": AgentTool(
        "fetch_url",
        "Fetch a public URL and return its text content. Use to read web pages, public documents, and online resources.",
        {**OBJECT, "properties": {"url": {"type": "string"}}, "required": ["url"]},
        _fetch_url,
    ),
    "web_search": AgentTool(
        "web_search",
        "Search the web and return top results with titles, snippets, and URLs. Use for finding current information, documentation, or references.",
        {
            **OBJECT,
            "properties": {
                "query": {"type": "string", "description": "Search query"},
                "num_results": {
                    "type": "integer",
                    "minimum": 1,
                    "maximum": 10,
                    "default": 5,
                },
            },
            "required": ["query"],
        },
        _web_search,
    ),
    "create_report": AgentTool(
        "create_report",
        "Save a generated document (report, comparison, summary, etc.) as a chat attachment accessible to the user. Use after reading source documents and generating new content. Supports markdown (.md) or PDF (.pdf) output.",
        {
            **OBJECT,
            "properties": {
                "session_id": {"type": "integer"},
                "content": {
                    "type": "string",
                    "description": "Full report/document content in markdown format",
                },
                "filename": {
                    "type": "string",
                    "description": "Filename including extension, e.g. comparison-report.md or comparison-report.pdf",
                },
                "format": {
                    "type": "string",
                    "enum": ["markdown", "pdf"],
                    "default": "markdown",
                    "description": "Output format: markdown or pdf",
                },
            },
            "required": ["session_id", "content"],
        },
        _create_report,
    ),
    "create_curriculum_version": AgentTool(
        "create_curriculum_version",
        "Create a named curriculum version snapshot (like a git commit). Use to checkpoint the curriculum state after a set of changes. Provide a descriptive name like 'feat: add CS201 lab experiments' or 'fix: correct credit hours for ECE301'.",
        {
            **OBJECT,
            "properties": {
                "name": {
                    "type": "string",
                    "description": "Descriptive version name (conventional commit style encouraged)",
                },
            },
            "required": ["name"],
        },
        _create_curriculum_version,
    ),
    "define_specialization": AgentTool(
        "define_specialization",
        "Create a specialization track (e.g. Machine Intelligence and Data Science) for a given semester. The letter (A, B, C...) is auto-assigned if omitted. Used to set up the elective specialization brackets.",
        {
            **OBJECT,
            "properties": {
                "semester": {
                    "type": "integer",
                    "minimum": 1,
                    "maximum": 8,
                    "description": "Semester the specialization applies to (5 or 6 for electives)",
                },
                "name": {
                    "type": "string",
                    "description": "Full specialization name, e.g. 'Machine Intelligence and Data Science (MIDS)'",
                },
                "letter": {
                    "type": "string",
                    "description": "Optional letter label (A, B, C). Auto-assigned if omitted.",
                },
                "key": {
                    "type": "string",
                    "description": "Optional short key (SCC, MIDS, CSCS). Derived from name if omitted.",
                },
                "academic_year": {
                    "type": "string",
                    "description": "Optional academic year batch, e.g. 2025-26",
                },
            },
            "required": ["semester", "name"],
        },
        _define_specialization,
    ),
    "list_specializations": AgentTool(
        "list_specializations",
        "List specialization track definitions, optionally filtered by semester. Use to discover which tracks exist before assigning electives.",
        {
            **OBJECT,
            "properties": {"semester": {"type": "integer", "minimum": 1, "maximum": 8}},
        },
        _list_specializations,
    ),
    "assign_elective_to_tracks": AgentTool(
        "assign_elective_to_tracks",
        "Categorize an elective course into one or more specialization tracks by their specialization_id. Also marks the course as an elective. Use after determining (e.g. via AI analysis) which tracks the course belongs to.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {"type": "integer"},
                "specialization_ids": {
                    "type": "array",
                    "items": {"type": "integer"},
                    "description": "Specialization track IDs to assign the elective to",
                },
            },
            "required": ["refined_id", "specialization_ids"],
        },
        _assign_elective_to_tracks,
    ),
    "remove_elective_from_tracks": AgentTool(
        "remove_elective_from_tracks",
        "Remove an elective course from one or more specialization tracks.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {"type": "integer"},
                "specialization_ids": {
                    "type": "array",
                    "items": {"type": "integer"},
                    "description": "Specialization track IDs to remove the elective from",
                },
            },
            "required": ["refined_id", "specialization_ids"],
        },
        _remove_elective_from_tracks,
    ),
    "get_course_assignments": AgentTool(
        "get_course_assignments",
        "Return which specialization tracks a course currently belongs to, including the track definitions.",
        {
            **OBJECT,
            "properties": {"refined_id": {"type": "integer"}},
            "required": ["refined_id"],
        },
        _get_course_assignments,
    ),
    "update_deterministic_fields": AgentTool(
        "update_deterministic_fields",
        "Create a reviewable draft that changes deterministic/protected fields (program, lecture_hours, tutorial_hours, practical_hours, self_study, credits, course_type). This is the ONLY way to modify those fields. The resulting draft is blocked until the user explicitly approves it in the Review panel. Always confirm with the user before calling this.",
        {
            **OBJECT,
            "properties": {
                "refined_id": {"type": "integer"},
                "fields": {
                    "type": "object",
                    "description": 'Protected fields to change, e.g. {"credits": 5, "lecture_hours": 4}',
                },
                "reason": {"type": "string"},
            },
            "required": ["refined_id", "fields"],
        },
        _update_deterministic_fields,
    ),
    "signal_done": AgentTool(
        "signal_done",
        "Signal that the agent has completed the user's request. Provide a concise summary of what was accomplished. This ends the agent's turn.",
        {
            **OBJECT,
            "properties": {
                "summary": {
                    "type": "string",
                    "description": "Brief summary of what was done, e.g. 'Created draft for CS201 adding Unit 5 on Graph Algorithms'",
                },
            },
            "required": ["summary"],
        },
        _signal_done,
    ),
    "create_spreadsheet": AgentTool(
        "create_spreadsheet",
        "Generate a CSV or Excel file from structured row data and save as a downloadable chat attachment. Provide column names and row objects.",
        {
            **OBJECT,
            "properties": {
                "session_id": {"type": "integer"},
                "columns": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Column names in display order",
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "object"},
                    "description": "Array of row objects, keys matching column names",
                },
                "filename": {
                    "type": "string",
                    "description": "Filename including extension, e.g. semester-3-courses.xlsx",
                },
                "format": {"type": "string", "enum": ["csv", "xlsx"], "default": "csv"},
            },
            "required": ["session_id", "columns", "rows"],
        },
        _create_spreadsheet,
    ),
    "get_version": AgentTool(
        "get_version",
        "Load a curriculum version snapshot with its metadata and course list. Use to inspect what a version contains before comparing or restoring.",
        {
            **OBJECT,
            "properties": {
                "version_id": {"type": "integer", "description": "Version snapshot ID"},
            },
            "required": ["version_id"],
        },
        _get_version,
    ),
    "diff_versions": AgentTool(
        "diff_versions",
        "Compare two curriculum version snapshots. Returns added, removed, and changed courses with per-course change percentages.",
        {
            **OBJECT,
            "properties": {
                "version_a": {
                    "type": "integer",
                    "description": "First version ID (base)",
                },
                "version_b": {
                    "type": "integer",
                    "description": "Second version ID (target)",
                },
            },
            "required": ["version_a", "version_b"],
        },
        _diff_versions,
    ),
    "get_curriculum_stats": AgentTool(
        "get_curriculum_stats",
        "Compute aggregate curriculum statistics: total courses, credits per semester, course type distribution, visible vs hidden counts. Optionally filter to one semester.",
        {
            **OBJECT,
            "properties": {
                "semester": {
                    "type": "integer",
                    "minimum": 1,
                    "maximum": 8,
                    "description": "Optional: stats for one semester only",
                },
            },
        },
        _get_curriculum_stats,
    ),
    "batch_read_courses": AgentTool(
        "batch_read_courses",
        "Read specific fields from multiple courses in one call. More efficient than calling get_course_fields multiple times. Returns an array of course field subsets.",
        {
            **OBJECT,
            "properties": {
                "refined_ids": {
                    "type": "array",
                    "items": {"type": "integer"},
                    "description": "List of course refined IDs to read",
                },
                "fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Field names to read for each course",
                },
            },
            "required": ["refined_ids", "fields"],
        },
        _batch_read_courses,
    ),
}
